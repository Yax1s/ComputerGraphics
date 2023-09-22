import openpyxl

path = r'D:\Bachelor of Science in Informatics and Computer Science\Year 3\Semester 2\Computer Graphics\Code\Repository\ComputerGraphics\Test Files.xlsx'

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(path)
 
# Define variable to read sheet
dataframe1 = dataframe.active

#Function to reverse a string
def reverse_string(x):
    return x[::-1]
 
# Iterate the loop to read the cell values
# for row in range(0, dataframe1.max_row):
#    for col in dataframe1.iter_cols(1, dataframe1.max_column):
#       print(col[row].value)

#Create email from name
email = "" #Variable to hold email
#Loop through all data, focusing on the name column
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(3, 3):
        name = col[row].value
        string_name = str(name)
        email = email + string_name[0] #Add the first letter of the name to the email
        reversed_string = reverse_string(string_name)
        last_name = ""
        #Loop through each reversed name until a space is reached then extract that substring
        for char in reversed_string:
            if char.isspace():
                true_name = reverse_string(last_name)
                email = email + true_name #Append the extracted name to the email
                email = email + '@email.com'
                print(str.lower(email))
                email = ""
                break
            else:
                last_name = last_name + str(char) #Append the character to the last_name variable
