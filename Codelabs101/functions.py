import openpyxl
import re

#Function to reverse a string
def reverse_string(x):
    return x[::-1]

#Function to create email from name
def create_emails(spreadsheet):
    emails = [] #Variable to hold all the emails
    email = "" #Variable to temporarily hold an email
    #Loop through all data, focusing on the name column
    for row in range(1, spreadsheet.max_row):
        for col in spreadsheet.iter_cols(3, 3):
            name = col[row].value
            string_name = str(name)
            email = email + string_name[0] #Add the first letter of the name to the email
            reversed_string = reverse_string(string_name)
            last_name = ""
            #Loop through each reversed name until a space is reached then extract that substring
            for char in reversed_string:
                if char.isspace():
                    true_name = reverse_string(last_name)
                    email = email + true_name #Append the extracted name to the email
                    email = email + '@email.com'
                    #Remove special characters
                    alphaEmail = re.sub("['@_!#$%^&*()<>?/\|}{~:]","",email)
                    emails.append(alphaEmail)
                    email = ""
                    alphaEmail = ""
                    break
                else:
                    last_name = last_name + str(char) #Append the character to the last_name variable

    return emails

#Function to check if the email addresses are unique
def unique_email(emails):
    unique = True
    for i in range(len(emails)):
        for j in range(len(emails)):
            if i != j:
                if emails[i] == emails[j]:
                    unique = False
                    print(emails[i])
    if unique:
        print("The email addresses are unique.")
    else:
        print("The above email address is repeated.")

#Function to return a list of all male students
def list_male(spreadsheet):
    males = []
    for row in range(1, spreadsheet.max_row):
        if (spreadsheet.cell(row, column = 6).value == 'M'):
            males.append(spreadsheet.cell(row, column = 3).value)
    return males

#Function to return a list of all female students
def list_female(spreadsheet):
    females = []
    for row in range(1, spreadsheet.max_row):
        if (spreadsheet.cell(row, column = 6).value == 'M'):
            females.append(spreadsheet.cell(row, column = 3).value)
    return females